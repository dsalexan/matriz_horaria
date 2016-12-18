import openpyxl
import json

FILE = "C:/Users/Danilo/Downloads/matriz_horaria_2017.1.xlsx"
TARGET = "C:/Users/Danilo/Projetos/matriz_horaria/matriz_output.json"
LABELS = {"first_label":"Horários",
          "trash": [None, "Docente"],
          "weekdays": {"Segunda":"seg", "Terça":"ter", "Quarta":"qua", "Quinta":"qui", "Sexta":"sex"}}
CONFIG = {"horizontal_fringe": 33,
          "pretty_json": False}


def detect_eof(sheet, index):
    return sheet["A" + str(index)].value is None and sheet["A" + str(index+1)].value is None and sheet["A" + str(index+2)].value is None

data = list()
xls = openpyxl.load_workbook(FILE)
sht = xls.active

i = 1
while sht["A{}".format(i)].value != LABELS['first_label']:
    i+=1
i+=1

""" i = primeira linha apos a etiqueta HORARIOS """


while not detect_eof(sht, i):
    if "A{}".format(i) in sht.merged_cells:
        termo = sht["A{}".format(i)].value.replace("\n", "")
        curso = termo[0:termo.find("Termo")].strip()
        termo = termo[termo.find("Termo"):].replace("Termo ", "").replace(" / ", "/").replace("/", ",")
        if termo.find("Termos ") != -1:
            termo = termo.replace("Termos ", "")
            termo = termo.replace(" e ", ",")

        if i == 26:
            asde1f32g = True

        mrg = [r for r in sht.merged_cell_ranges if r.find("A{}:".format(i)) != -1][0]
        for j in range(0, len(sht[mrg])):
            if not sht["B{}".format(i+j)].value in LABELS['trash']:
                subjects = {"seg": [],
                            "ter": [],
                            "qua": [],
                            "qui": [],
                            "sex": []}
                for k in range(3, CONFIG["horizontal_fringe"] + 1):
                    if not sht.cell(column=k, row=2).value is None:
                        dia = sht.cell(column=k, row=2).value.strip()

                    if not sht.cell(column=k, row=i+j).value is None:
                        dta = {"subject": sht.cell(column=k, row=i+j).value.replace("\n", "").replace(" Turma ", "Turma ").replace(" Turma", "Turma ").replace("Turma ", "Turma").replace("Turma", " ").replace("-", "").replace("  ", " "), "professor": sht.cell(column=k, row=i+j+1).value}

                        if k == 11 and i+j == 28:
                            asdasde = True

                        subjects[LABELS["weekdays"][dia]].append((dta.copy()))

                data.append({"cell": "A{}".format(i+j),
                             "course": curso,
                             "term": termo.split(','),
                             "time": sht["B{}".format(i+j)].value.replace(" ", ""),
                             "week": subjects})

        i += len(sht[mrg]) - 1

    i+=1

with open(TARGET, "wb") as trg:
    if CONFIG['pretty_json']:
        compressed = (json.dumps(data, ensure_ascii=False, sort_keys=False, indent=3, separators=(',', ': ')))
    else:
        compressed = (json.dumps(data, ensure_ascii=False, sort_keys=False, indent=None, separators=None))

    trg.write(compressed.encode("utf-8"))

print(compressed)
""""""